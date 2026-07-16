#!/usr/bin/env python3
"""Build the word_art wordlist CSV from public sources.

Sources:
- Brysbaert, Warriner & Kuperman (2014) concreteness ratings for 40k
  English lemmas -- Springer supplementary xlsx.
- Google Quick, Draw! categories -- 345 pre-vetted drawable concepts.
- Hand-curated seed of abstract nouns (checked in under
  scripts/word_art_wordlist/abstracts_seed.txt).

Output: kaggle_environments/envs/word_art/words.csv, columns
    word,category,tier,source

The CSV is the runtime source of truth. This script downloads the
external datasets to scripts/.cache/word_art_wordlist/ (not checked in)
and reuses them on subsequent runs unless --refresh is passed.

Usage:
    UV_FROZEN=1 uv run --with openpyxl --with nltk \\
        scripts/build_word_art_wordlist.py [--refresh]

After regenerating: run the matcher-trim guard test --
    UV_FROZEN=1 uv run pytest tests/envs/word_art/test_word_art.py -k trimmed
The plural-matching tables in word_art.py are CSV-scoped; the guard test
names any entries to add or drop after a CSV shift.
"""
from __future__ import annotations

import argparse
import csv
import sys
import urllib.request
from collections import Counter
from pathlib import Path

SCRIPT_DIR = Path(__file__).resolve().parent
REPO_ROOT = SCRIPT_DIR.parent
CACHE_DIR = SCRIPT_DIR / ".cache" / "word_art_wordlist"
SEEDS_DIR = SCRIPT_DIR / "word_art_wordlist"
OUT_PATH = REPO_ROOT / "kaggle_environments" / "envs" / "word_art" / "words.csv"

# The old crr.ugent.be paper URL 404s (checked 2026-07). The original
# supplementary material lives under the Springer DOI landing page and
# is stable.
BRYSBAERT_URL = (
    "https://static-content.springer.com/esm/"
    "art%3A10.3758%2Fs13428-013-0403-5/MediaObjects/"
    "13428_2013_403_MOESM1_ESM.xlsx"
)
QUICKDRAW_URL = (
    "https://raw.githubusercontent.com/"
    "googlecreativelab/quickdraw-dataset/master/categories.txt"
)

# --- Filter thresholds ------------------------------------------------------
# Nouns: high concreteness, strong rater agreement, broadly known, and a
# recognisability sanity floor on SUBTLEX frequency (raw count in the
# 51M-word SUBTLEX-US corpus). Upper SUBTLEX bound trims function-word-ish
# ultra-frequent lemmas that leak past the noun-POS filter.
NOUN_CONC_MIN = 4.0
NOUN_CONC_SD_MAX = 1.0
NOUN_PERCENT_KNOWN_MIN = 0.90
NOUN_SUBTLEX_MIN = 5
NOUN_SUBTLEX_MAX = 200_000

# Verbs: lower concreteness bar (action verbs like JUGGLE rate 3.5-4.5,
# not 4.5+), and a stricter frequency floor to keep the pool grounded.
# Top-N by SUBTLEX gives a fixed-size, well-known verb list.
VERB_CONC_MIN = 3.5
VERB_SUBTLEX_MIN = 20
VERB_SUBTLEX_MAX = 200_000
VERB_TOP_N = 400

# Heuristic tier boundaries for nouns. QuickDraw membership auto-promotes
# to easy because those categories have been vetted as drawable in ~20s
# by humans, which is a strong prior even for lower-concreteness entries.
NOUN_EASY_CONC = 4.7
NOUN_MEDIUM_CONC = 4.3


# Category taxonomy used by the human-curated blocklist. The blocklist
# file itself is gitignored so the sensitive terms never enter git
# history; this list preserves the *intent* -- what a fresh maintainer
# needs to aggressively block when reconstructing the file. When in
# doubt, err toward blocking. Each entry corresponds to a `## Group`
# header in scripts/word_art_wordlist/blocklist.txt.
BLOCKLIST_CATEGORIES = (
    "profanity / vulgar",
    "sexual / anatomical / body-part slang",
    "pregnancy / reproductive / genitourinary specialists",
    "clinical excreta / bodily fluids / gross secretions",
    "clinical / awkward internal anatomy",
    "slurs / slur-adjacent",
    "drug-adjacent (illegal drugs, pharmaceuticals, paraphernalia)",
    "historically loaded / racialized-violence connotations",
    "outdated / exoticizing / dismissive identity labels",
    "family / kinship labels (all gendered relationship terms)",
    "dog breeds (blocked as a class)",
    "brand / trademarked references",
    "gendered occupation labels (prefer neutral forms)",
    "informal gendered person-labels",
    "politically-charged / civil-unrest / political-violence / "
        "military ranks / law enforcement",
    "grim / illness / mortality",
    "occult / dark supernatural",
    "body / appearance shaming (incl. form-fitting clothing)",
    "animal-cruelty coded / meat products",
    "religion (except CHURCH as neutral place-of-worship)",
    "gambling / vice",
    "alcohol (except broadly acceptable dual-use: DRINK, BAR, WINE, etc.)",
    "tobacco / smoking",
    "modern firearms + explosives (historical/fantasy weapons kept)",
)


def _load_blocklist(path: Path) -> frozenset[str]:
    """Read the human-authored blocklist from a plain-text file.

    Line-oriented format: ``## Group`` headers and any ``#``-prefixed
    line are comments; blank lines skipped; every other line is one
    uppercase word to block.

    The file at scripts/word_art_wordlist/blocklist.txt is gitignored --
    it enumerates sensitive terms we don't want in git history. On a
    fresh clone this returns an empty set with a warning; get the file
    from the wordlist maintainer to reproduce the shipped CSV. See
    BLOCKLIST_CATEGORIES above for the taxonomy of what to block.
    """
    if not path.exists():
        print(
            f"WARNING: blocklist not found at {path}. Proceeding with an "
            "EMPTY blocklist -- the generated CSV will NOT be filtered for "
            "sensitive terms. See BLOCKLIST_CATEGORIES in this file for the "
            "taxonomy to reconstruct, and _load_blocklist for the format.",
            file=sys.stderr,
        )
        return frozenset()
    out: set[str] = set()
    with open(path) as f:
        for line in f:
            s = line.strip()
            if not s or s.startswith("#"):
                continue
            out.add(s.upper())
    return frozenset(out)


# Words removed from every source before writing the CSV. See
# scripts/word_art_wordlist/blocklist.txt for the categorized list and
# the reasoning per group.
BLOCKLIST = _load_blocklist(SEEDS_DIR / "blocklist.txt")


def _fetch(url: str, dest: Path, refresh: bool) -> None:
    if dest.exists() and not refresh:
        return
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    print(f"Downloading {url}")
    urllib.request.urlretrieve(url, dest)


def _load_brysbaert(xlsx_path: Path) -> list[dict]:
    """Return one dict per unigram entry.

    The xlsx's 'Bigram' column is 1 for multi-word phrases; we drop those
    here since v1 doesn't ship compounds. Every remaining row becomes a
    candidate for either the noun or verb pool depending on POS.
    """
    from openpyxl import load_workbook

    wb = load_workbook(str(xlsx_path), read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = next(rows)
    idx = {name: i for i, name in enumerate(header)}
    out: list[dict] = []
    for r in rows:
        word = r[idx["Word"]]
        bigram = r[idx["Bigram"]]
        if word is None or bigram != 0:
            continue
        out.append(
            {
                "word": word,
                "conc_m": r[idx["Conc.M"]] or 0.0,
                "conc_sd": r[idx["Conc.SD"]] or 0.0,
                "percent_known": r[idx["Percent_known"]] or 0.0,
                "subtlex": r[idx["SUBTLEX"]] or 0,
            }
        )
    return out


def _load_quickdraw(txt_path: Path) -> list[str]:
    with open(txt_path) as f:
        return [line.strip() for line in f if line.strip()]


def _load_seed(name: str) -> list[str]:
    """Load a hand-curated seed list. Blank lines and `#`-comments ignored."""
    p = SEEDS_DIR / f"{name}_seed.txt"
    out: list[str] = []
    with open(p) as f:
        for line in f:
            s = line.strip()
            if not s or s.startswith("#"):
                continue
            out.append(s)
    return out


def _looks_ok_alpha(word: str) -> bool:
    """Accept only single-token, ASCII, alphabetic lemmas.

    Brysbaert lemmas are already lowercase, but hyphenated forms and
    numeric tokens ('7up') slip through the concreteness filter. Also
    drop anything starting uppercase as a defensive proper-noun filter.
    """
    if not word:
        return False
    if not word.isascii() or not word.isalpha():
        return False
    if " " in word or "-" in word or "_" in word:
        return False
    if word[0].isupper():
        return False
    return True


def _wn_pos(word: str, wn) -> str | None:
    """Dominant POS between noun and verb by synset count, or None if neither.

    Ties break to noun -- most concrete lemmas that could go either way
    (PAINT, DRAW, ANCHOR) are more useful drawn as nouns.
    """
    n = len(wn.synsets(word, pos="n"))
    v = len(wn.synsets(word, pos="v"))
    if n == 0 and v == 0:
        return None
    if n >= v:
        return "noun"
    return "verb"


def _drop_redundant_inflections(rows: list[dict], wn) -> tuple[list[dict], list[str]]:
    """Drop any row whose morphy-lemma is a different word already in the CSV.

    Handles noun plurals (BATS -> BAT), verb inflections including same-length
    irregulars (WROTE -> WRITE, ATE -> EAT, SAT -> SIT, WENT -> GO), and
    gerunds/participles (BEATING -> BEAT, BROKEN -> BREAK, BURNT -> BURN)
    that Brysbaert's lemmatisation leaks through. Only drops when the base
    form is also present, so no concept is lost -- true plurale tantum
    (JEANS, CLOTHES) whose singular is absent stay.

    Note: no length check -- earlier version used ``len(lemma) >= len(w)``
    as a defensive guard, which incorrectly kept same-length irregulars
    like WROTE (5 chars, lemma WRITE also 5 chars).
    """
    all_words = {r["word"] for r in rows}
    kept: list[dict] = []
    dropped: list[str] = []
    for r in rows:
        w = r["word"]
        pos = wn.VERB if r["category"] == "verb" else wn.NOUN
        lemma = wn.morphy(w.lower(), pos)
        if lemma is None:
            kept.append(r)
            continue
        lemma_up = lemma.upper()
        if lemma_up == w:
            kept.append(r)
            continue
        if lemma_up in all_words:
            dropped.append(w)
            continue
        kept.append(r)
    return kept, dropped


def _drop_synonyms(rows: list[dict], wn) -> tuple[list[dict], list[tuple[str, str]]]:
    """Collapse WordNet-synonym clusters to a single representative per group.

    Group CSV words by their *primary synset ID* (POS = row category mapped
    to NOUN/VERB, first synset from wn.synsets). Two words in the same
    primary-synset group refer to the same concept -- e.g. SOFA and COUCH
    both have primary synset ``sofa.n.01``, so drawing one and guessing
    the other should not count as a loss. Within each such group, keep the
    alphabetically-first word at the easiest tier and drop the rest.

    Strict same-synset grouping (rather than lemma-based union-find) avoids
    transitive false merges: earlier versions chained AVIATOR->BANKNOTE via
    a shared lemma across different senses.

    Returns (kept_rows, list_of (dropped_word, kept_representative) pairs).
    """
    csv_by_word = {r["word"]: r for r in rows}

    # Group by primary synset ID. Two words with the *same* primary synset
    # refer to the same concept (SOFA and COUCH both -> sofa.n.01, MOVIE
    # and FILM both -> movie.n.01, AUTOMOBILE and CAR both -> car.n.01).
    # We deliberately do NOT relax to "shares any synset" or "primary of
    # one appears in the other's synset list": those relaxations bring
    # transitive false merges via overly-broad secondary synsets (e.g. a
    # money-slang synset with a dozen lemmas can chain AVIATOR to BANKNOTE
    # via BILL). Cost of being strict: genuine near-synonym pairs like
    # JAIL/PRISON, HAT/CAP, CAB/TAXI keep both members. That's acceptable
    # -- both are semantically distinct enough in WordNet, and no false
    # merges is worth some remaining redundancy.
    groups: dict[str, list[str]] = {}
    for r in rows:
        w = r["word"]
        pos = wn.VERB if r["category"] == "verb" else wn.NOUN
        synsets = wn.synsets(w.lower(), pos=pos)
        if not synsets:
            continue
        groups.setdefault(synsets[0].name(), []).append(w)

    tier_rank = {"easy": 0, "medium": 1, "hard": 2}
    dropped_pairs: list[tuple[str, str]] = []
    dropped_set: set[str] = set()
    for members in groups.values():
        if len(members) <= 1:
            continue
        # Preference order (all "lower is better"):
        #   (a) original-list membership -- if the pre-build hand-curated
        #       list picked this word, respect that curation
        #   (b) easier tier
        #   (c) shorter word (proxy for more common / more natural drawing
        #       target -- CAR over AUTOMOBILE, CHIMP over CHIMPANZEE)
        #   (d) alphabetical as final tiebreaker
        keeper = min(
            members,
            key=lambda w: (
                0 if "original" in csv_by_word[w]["source"] else 1,
                tier_rank[csv_by_word[w]["tier"]],
                len(w),
                w,
            ),
        )
        for m in members:
            if m != keeper:
                dropped_pairs.append((m, keeper))
                dropped_set.add(m)

    kept = [r for r in rows if r["word"] not in dropped_set]
    return kept, sorted(dropped_pairs)


def _tier_for_noun(conc_m: float, in_quickdraw: bool) -> str:
    if conc_m >= NOUN_EASY_CONC or in_quickdraw:
        return "easy"
    if conc_m >= NOUN_MEDIUM_CONC:
        return "medium"
    return "hard"


def _ensure_wordnet():
    """Load WordNet, downloading the corpus on first run if needed."""
    import nltk
    from nltk.corpus import wordnet as wn

    try:
        wn.synsets("dog")
    except LookupError:
        print("Downloading NLTK WordNet corpus (~30MB, one-time) ...")
        nltk.download("wordnet", quiet=True)
        # Re-import after download.
        from nltk.corpus import wordnet as wn2

        return wn2
    return wn


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--refresh", action="store_true", help="Force re-download of source datasets.")
    args = ap.parse_args()

    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    brysbaert_path = CACHE_DIR / "brysbaert.xlsx"
    quickdraw_path = CACHE_DIR / "quickdraw_categories.txt"
    _fetch(BRYSBAERT_URL, brysbaert_path, args.refresh)
    _fetch(QUICKDRAW_URL, quickdraw_path, args.refresh)

    wn = _ensure_wordnet()
    bry = _load_brysbaert(brysbaert_path)
    quickdraw_cats = _load_quickdraw(quickdraw_path)
    quickdraw_upper = {c.upper() for c in quickdraw_cats}
    abstracts_seed = _load_seed("abstracts")

    rows: list[dict] = []
    seen: set[str] = set()
    blocked_hits: list[str] = []  # collected for the end-of-run report

    def _blocked(up: str) -> bool:
        if up in BLOCKLIST:
            blocked_hits.append(up)
            return True
        return False

    # --- Nouns from Brysbaert -------------------------------------------
    for entry in bry:
        w = entry["word"]
        if not _looks_ok_alpha(w):
            continue
        if entry["conc_m"] < NOUN_CONC_MIN:
            continue
        if entry["conc_sd"] > NOUN_CONC_SD_MAX:
            continue
        if entry["percent_known"] < NOUN_PERCENT_KNOWN_MIN:
            continue
        if not (NOUN_SUBTLEX_MIN <= entry["subtlex"] <= NOUN_SUBTLEX_MAX):
            continue
        if _wn_pos(w, wn) != "noun":
            continue
        up = w.upper()
        if up in seen or _blocked(up):
            continue
        seen.add(up)
        in_qd = up in quickdraw_upper
        rows.append(
            {
                "word": up,
                "category": "noun",
                "tier": _tier_for_noun(entry["conc_m"], in_qd),
                "source": "brysbaert+quickdraw" if in_qd else "brysbaert",
            }
        )

    # --- QuickDraw-only nouns -------------------------------------------
    # QuickDraw supplies iconic drawable concepts (~single words only for
    # v1) missing from Brysbaert or filtered out by our stricter thresholds.
    # These are always tier=easy because QuickDraw already vetted them for
    # human draw-ability.
    for cat in quickdraw_cats:
        if " " in cat or "-" in cat:
            continue
        up = cat.upper()
        if up in seen or _blocked(up):
            continue
        seen.add(up)
        rows.append(
            {"word": up, "category": "noun", "tier": "easy", "source": "quickdraw"}
        )

    # --- Verbs from Brysbaert -------------------------------------------
    verb_candidates: list[tuple[int, str]] = []
    for entry in bry:
        w = entry["word"]
        if not _looks_ok_alpha(w):
            continue
        if entry["conc_m"] < VERB_CONC_MIN:
            continue
        if not (VERB_SUBTLEX_MIN <= entry["subtlex"] <= VERB_SUBTLEX_MAX):
            continue
        if _wn_pos(w, wn) != "verb":
            continue
        up = w.upper()
        if up in seen or _blocked(up):
            continue
        verb_candidates.append((entry["subtlex"], up))
    verb_candidates.sort(reverse=True)
    for _, up in verb_candidates[:VERB_TOP_N]:
        seen.add(up)
        rows.append({"word": up, "category": "verb", "tier": "hard", "source": "brysbaert"})

    # --- Abstracts from seed --------------------------------------------
    for w in abstracts_seed:
        up = w.upper()
        if up in seen or _blocked(up):
            continue
        seen.add(up)
        rows.append({"word": up, "category": "abstract", "tier": "hard", "source": "seed"})

    # --- Originals allowlist --------------------------------------------
    # The pre-build hand-curated 368-word wordlist. Any of these that
    # aren't already added get force-included, bypassing the concreteness /
    # SD / POS-dominance filters that would otherwise drop legitimate
    # drawable objects like ABACUS, BEETLE, PALETTE. Each is a
    # human-vetted Pictionary target, so treating them as authoritative
    # for drawability. Still respects BLOCKLIST -- an original word
    # that's on the blocklist stays out.
    #
    # For originals already added by another source, we tag their source
    # with "+original" so the synonym-cluster tiebreaker in _drop_synonyms
    # can prefer originals even when both members were pulled from Brysbaert.
    originals_seed = _load_seed("originals")
    rows_by_word = {r["word"]: r for r in rows}
    for w in originals_seed:
        up = w.upper()
        if _blocked(up):
            continue
        if up in seen:
            existing = rows_by_word[up]
            if "original" not in existing["source"]:
                existing["source"] = f"{existing['source']}+original"
            continue
        seen.add(up)
        rows.append({"word": up, "category": "noun", "tier": "easy", "source": "original"})

    # Drop plurals and other non-lemma inflections whose base form is also
    # present -- pure redundancy with no information loss. Plurale tantum
    # (JEANS, CLOTHES, INTESTINES) whose singular is absent are kept.
    rows, dropped_inflections = _drop_redundant_inflections(rows, wn)

    # Collapse WordNet-synonym clusters (SOFA/COUCH, CAB/TAXI, JAIL/PRISON)
    # so a correct-concept drawing can't be marked wrong just because the
    # guesser picked the peer synonym. See _drop_synonyms.
    rows, dropped_synonym_pairs = _drop_synonyms(rows, wn)

    # Sort so the diff is stable across re-runs when the source data changes.
    rows.sort(key=lambda r: (r["word"], r["category"]))

    cat_counts = Counter(r["category"] for r in rows)
    tier_counts = Counter(r["tier"] for r in rows)
    print()
    print(f"Total words: {len(rows)}")
    print(f"By category: {dict(cat_counts)}")
    print(f"By tier:     {dict(tier_counts)}")
    if blocked_hits:
        print(f"Blocked (removed by BLOCKLIST): {sorted(set(blocked_hits))}")
    if dropped_inflections:
        print(f"Dropped {len(dropped_inflections)} redundant plurals/inflections; "
              f"first 20: {sorted(dropped_inflections)[:20]}")
    if dropped_synonym_pairs:
        print(f"Dropped {len(dropped_synonym_pairs)} synonym-cluster members; "
              f"first 50 (dropped -> kept):")
        for d, k in dropped_synonym_pairs[:50]:
            print(f"  {d} -> {k}")

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT_PATH, "w", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=["word", "category", "tier", "source"])
        writer.writeheader()
        writer.writerows(rows)
    print(f"Wrote {OUT_PATH}")


if __name__ == "__main__":
    main()
